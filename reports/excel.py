from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def render_excel_report(*, title: str, rows: list[dict], filename: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append([title])
    ws.append([])

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        for row in rows:
            ws.append([str(row.get(h, "")) for h in headers])

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_length:
                    max_length = len(val)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    else:
        ws.append(["Sin datos para mostrar"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
